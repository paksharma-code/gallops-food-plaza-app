"""
Quick regression for Gallops Food Plaza backend after UI-fix batch.

Focus:
 1. POST /api/offer-claims flexible DOB/Anniversary parsing (separators + ISO + empty + garbage)
    - Stored dob/anniversary must be ISO (YYYY-MM-DD) or empty
 2. GET /api/admin/offer-claims/export?token=<jwt>
    - Headers = prior 11 columns
    - DOB/Anniversary columns rendered as DD-MM-YYYY
    - Mobile column formatted "+91XXXXXXXXXX"
 3. Regression:
    - POST /api/offer-claims with outlet_id → whatsapp_link non-empty + wa.me/<digits>?text=
    - is_offer_of_the_day round-trip (POST /api/offers)
    - Smoke GET /api/plazas, /api/outlets, /api/offers, /api/menu
 4. Cleanup: DELETE /api/admin/offer-claims/{id}
"""
import os
import re
import sys
from io import BytesIO

import requests
from openpyxl import load_workbook

BASE = os.environ.get(
    "TEST_BASE_URL",
    "https://gallops-reserve-dine.preview.emergentagent.com",
).rstrip("/")
API = f"{BASE}/api"

ADMIN_EMAIL = "admin@gallops.com"
ADMIN_PASSWORD = "admin123"

results = {"pass": [], "fail": []}
group_results = {}  # group name -> list of (name, ok, msg)
cleanup_claim_ids = []
cleanup_offer_ids = []
admin_token = None


def log(group, name, ok, extra=""):
    symbol = "PASS" if ok else "FAIL"
    msg = f"{symbol}  [{group}] {name}"
    if extra:
        msg += f" — {extra}"
    print(msg)
    (results["pass"] if ok else results["fail"]).append(f"[{group}] {name}{' — ' + extra if extra else ''}")
    group_results.setdefault(group, []).append((name, ok, extra))


def auth_headers():
    return {"Authorization": f"Bearer {admin_token}"}


def login():
    global admin_token
    r = requests.post(f"{API}/auth/login", json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD}, timeout=30)
    assert r.status_code == 200, f"login failed: {r.status_code} {r.text}"
    admin_token = r.json()["token"]
    print(f"Logged in, token len={len(admin_token)}")


def post_claim(payload, label, expected_dob, expected_anniv, group="1. Date parsing"):
    """POST /api/offer-claims and verify storage via GET admin."""
    r = requests.post(f"{API}/offer-claims", json=payload, timeout=30)
    if r.status_code != 200:
        log(group, f"POST claim {label}", False, f"status {r.status_code} body={r.text[:200]}")
        return None
    data = r.json()
    cid = data.get("id")
    if not cid:
        log(group, f"POST claim {label}", False, "no id in response")
        return None
    cleanup_claim_ids.append(cid)

    # Now fetch via admin list and confirm stored dob / anniversary
    rl = requests.get(
        f"{API}/admin/offer-claims",
        headers=auth_headers(),
        params={"mobile": payload["mobile"]},
        timeout=30,
    )
    if rl.status_code != 200:
        log(group, f"GET admin claim {label}", False, f"status {rl.status_code}")
        return cid
    items = rl.json()
    match = next((it for it in items if it.get("id") == cid), None)
    if not match:
        log(group, f"GET admin claim {label}", False, "claim not found in admin list")
        return cid
    actual_dob = match.get("dob", "")
    actual_anniv = match.get("anniversary", "")
    if actual_dob == expected_dob and actual_anniv == expected_anniv:
        log(group, f"DOB/Anniversary normalised for {label}", True,
            f"stored dob={actual_dob!r} anniversary={actual_anniv!r}")
    else:
        log(group, f"DOB/Anniversary normalised for {label}", False,
            f"expected dob={expected_dob!r}/anniv={expected_anniv!r}, got dob={actual_dob!r}/anniv={actual_anniv!r}")
    return cid


def group1_date_parsing():
    print("\n=== Group 1: Flexible DOB/Anniversary parsing ===")
    cases = [
        ("a dash DD-MM-YYYY",      {"name": "QA User", "mobile": "9876543210", "dob": "15-05-1990", "anniversary": "10-12-2015"}, "1990-05-15", "2015-12-10"),
        ("b slash DD/MM/YYYY",     {"name": "QA User", "mobile": "9876543211", "dob": "15/05/1990", "anniversary": "10/12/2015"}, "1990-05-15", "2015-12-10"),
        ("c dot DD.MM.YYYY",       {"name": "QA User", "mobile": "9876543212", "dob": "15.05.1990", "anniversary": "10.12.2015"}, "1990-05-15", "2015-12-10"),
        ("d ISO passthrough",      {"name": "QA User", "mobile": "9876543213", "dob": "1990-05-15", "anniversary": "2015-12-10"}, "1990-05-15", "2015-12-10"),
        ("e empty strings",        {"name": "QA User", "mobile": "9876543214", "dob": "", "anniversary": ""}, "", ""),
        ("f garbage values",       {"name": "QA User", "mobile": "9876543215", "dob": "garbage", "anniversary": "garbage"}, "", ""),
    ]
    for label, payload, exp_dob, exp_anniv in cases:
        post_claim(payload, label, exp_dob, exp_anniv)


def group2_xlsx_export():
    print("\n=== Group 2: XLSX export formatting ===")
    group = "2. XLSX export"
    r = requests.get(
        f"{API}/admin/offer-claims/export",
        params={"token": admin_token},
        timeout=60,
    )
    if r.status_code != 200:
        log(group, "GET export", False, f"status {r.status_code} body={r.text[:200]}")
        return
    ctype = r.headers.get("Content-Type", "")
    if "spreadsheetml.sheet" not in ctype:
        log(group, "Content-Type xlsx", False, f"got {ctype}")
    else:
        log(group, "Content-Type xlsx", True, ctype)

    wb = load_workbook(BytesIO(r.content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        log(group, "xlsx has rows", False, "0 rows")
        return
    header = list(rows[0])
    expected_headers = [
        "Date of Entry (IST)",
        "Name",
        "Mobile Number",
        "DOB",
        "Anniversary",
        "Plaza",
        "Outlet",
        "Offer",
        "Token",
        "Valid Till (IST)",
        "Offer Availed",
    ]
    if header == expected_headers:
        log(group, "11-column header intact", True, f"{len(header)} cols")
    else:
        log(group, "11-column header intact", False, f"got={header}")

    # Find QA rows (mobile starts with +91987654321)
    qa_rows = [r for r in rows[1:] if r and len(r) >= 5 and str(r[2] or "").startswith("+91987654321")]
    log(group, "QA rows present in export", len(qa_rows) >= 6, f"{len(qa_rows)} rows found")

    # DOB / Anniversary DD-MM-YYYY check across QA rows.
    dd_mm_yyyy = re.compile(r"^\d{2}-\d{2}-\d{4}$")
    dob_format_ok = True
    anniv_format_ok = True
    dob_samples = []
    anniv_samples = []
    for row in qa_rows:
        dob_val = str(row[3] or "")
        anniv_val = str(row[4] or "")
        dob_samples.append(dob_val)
        anniv_samples.append(anniv_val)
        if dob_val and not dd_mm_yyyy.match(dob_val):
            dob_format_ok = False
        if anniv_val and not dd_mm_yyyy.match(anniv_val):
            anniv_format_ok = False

    log(group, "DOB column DD-MM-YYYY", dob_format_ok, f"samples={dob_samples}")
    log(group, "Anniversary column DD-MM-YYYY", anniv_format_ok, f"samples={anniv_samples}")

    # Specifically check at least one row has "15-05-1990" for DOB
    has_15_05_1990 = any("15-05-1990" == str(r[3] or "") for r in qa_rows)
    has_10_12_2015 = any("10-12-2015" == str(r[4] or "") for r in qa_rows)
    log(group, "Found DOB=15-05-1990 in export", has_15_05_1990)
    log(group, "Found Anniversary=10-12-2015 in export", has_10_12_2015)

    # Mobile column formatted as "+91XXXXXXXXXX" (13 chars total)
    mobile_fmt_ok = all(
        re.match(r"^\+91\d{10}$", str(r[2] or ""))
        for r in qa_rows
    )
    log(group, "Mobile column +91XXXXXXXXXX", mobile_fmt_ok, f"sample={qa_rows[0][2] if qa_rows else None}")


def group3_regression():
    print("\n=== Group 3: Regression checks ===")
    group = "3. Regression"

    # 3a: claim with outlet_id → whatsapp_link well-formed
    outlets = requests.get(f"{API}/outlets", timeout=30).json()
    if not outlets:
        log(group, "fetch outlets", False, "no outlets")
        return
    # pick any outlet with a mobile
    outlet = next((o for o in outlets if o.get("mobile")), outlets[0])
    payload = {
        "name": "Regression User",
        "mobile": "9876500001",
        "outlet_id": outlet["id"],
    }
    r = requests.post(f"{API}/offer-claims", json=payload, timeout=30)
    if r.status_code == 200:
        body = r.json()
        wa = body.get("whatsapp_link", "")
        cleanup_claim_ids.append(body["id"])
        ok = bool(wa) and wa.startswith("https://wa.me/") and "?text=" in wa
        # also check the digits part is all digits
        if ok:
            m = re.match(r"^https://wa\.me/(\d+)\?text=", wa)
            ok = bool(m) and len(m.group(1)) >= 10
        log(group, "whatsapp_link well-formed (wa.me/<digits>?text=...)", ok, wa[:120])
    else:
        log(group, "POST claim with outlet_id", False, f"status {r.status_code}")

    # 3b: is_offer_of_the_day round-trip
    offer_payload = {
        "title": "QA Regression Offer",
        "description": "Temporary regression offer",
        "is_active": True,
        "is_offer_of_the_day": True,
    }
    rc = requests.post(f"{API}/offers", headers=auth_headers(), json=offer_payload, timeout=30)
    if rc.status_code != 200:
        log(group, "POST offer with is_offer_of_the_day", False, f"status {rc.status_code} body={rc.text[:200]}")
    else:
        oid = rc.json()["id"]
        cleanup_offer_ids.append(oid)
        # fetch list
        rl = requests.get(f"{API}/offers", timeout=30).json()
        match = next((o for o in rl if o.get("id") == oid), None)
        ok = match is not None and match.get("is_offer_of_the_day") is True
        log(group, "is_offer_of_the_day=true persists on GET", ok, f"got={match.get('is_offer_of_the_day') if match else None}")

    # 3c: Smoke GETs
    for path, label in [("/plazas", "GET /api/plazas"), ("/outlets", "GET /api/outlets"),
                        ("/offers", "GET /api/offers"), ("/menu", "GET /api/menu")]:
        r = requests.get(f"{API}{path}", timeout=60)
        ok = r.status_code == 200 and isinstance(r.json(), list)
        count = len(r.json()) if ok else 0
        log(group, label, ok, f"status={r.status_code} count={count}")


def group4_cleanup():
    print("\n=== Group 4: Cleanup ===")
    group = "4. Cleanup"
    deleted_claims = 0
    for cid in list(set(cleanup_claim_ids)):
        r = requests.delete(f"{API}/admin/offer-claims/{cid}", headers=auth_headers(), timeout=30)
        if r.status_code == 200:
            deleted_claims += 1
        else:
            log(group, f"DELETE claim {cid}", False, f"status {r.status_code}")
    log(group, f"Deleted {deleted_claims} QA claims", deleted_claims == len(set(cleanup_claim_ids)),
        f"expected {len(set(cleanup_claim_ids))}, got {deleted_claims}")

    # Also cleanup test offer
    for oid in cleanup_offer_ids:
        r = requests.delete(f"{API}/offers/{oid}", headers=auth_headers(), timeout=30)


def main():
    print(f"BASE = {BASE}")
    login()
    group1_date_parsing()
    group2_xlsx_export()
    group3_regression()
    group4_cleanup()

    print("\n" + "=" * 70)
    print(f"TOTAL PASS: {len(results['pass'])}")
    print(f"TOTAL FAIL: {len(results['fail'])}")
    if results["fail"]:
        print("\nFailures:")
        for f in results["fail"]:
            print(" - " + f)
    print("\nPer-group summary:")
    for g, rows in group_results.items():
        passed = sum(1 for _, ok, _ in rows if ok)
        total = len(rows)
        print(f"  {g}: {passed}/{total} passed")

    sys.exit(0 if not results["fail"] else 1)


if __name__ == "__main__":
    main()
