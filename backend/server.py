from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

import os
import uuid
import logging
import asyncio
from datetime import datetime, timezone, timedelta
from typing import List, Optional

import bcrypt
import jwt
from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, Query
from fastapi.staticfiles import StaticFiles
from starlette.middleware.cors import CORSMiddleware
from starlette.middleware.gzip import GZipMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, EmailStr, Field


# ----------------- Setup -----------------
MONGO_URL = os.environ["MONGO_URL"]
DB_NAME = os.environ["DB_NAME"]
JWT_SECRET = os.environ["JWT_SECRET"]
JWT_ALGORITHM = "HS256"
ADMIN_EMAIL = os.environ.get("ADMIN_EMAIL", "admin@gallops.com")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "admin123")

client = AsyncIOMotorClient(MONGO_URL)
db = client[DB_NAME]

app = FastAPI(title="Gallops Food Plaza API")
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s")
logger = logging.getLogger("gallops")


# ----------------- Models -----------------
class LoginRequest(BaseModel):
    email: EmailStr
    password: str


class LoginResponse(BaseModel):
    token: str
    user: dict


class ChangePasswordRequest(BaseModel):
    current_password: str
    new_password: str


class Plaza(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    city: Optional[str] = ""
    status: str = "operational"  # "operational" | "upcoming"
    is_head_office: bool = False
    description: Optional[str] = ""
    address: Optional[str] = ""
    image: Optional[str] = None
    image2: Optional[str] = None  # extra gallery image #1 (optional)
    image3: Optional[str] = None  # extra gallery image #2 (optional)
    gallery: Optional[List[str]] = []
    google_maps_url: Optional[str] = ""
    contact_phone: Optional[str] = ""
    whatsapp_number: Optional[str] = ""  # plaza-level WhatsApp for offer claims fallback
    expected_opening: Optional[str] = ""
    order_index: int = 0
    is_offers_enabled: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class PlazaCreate(BaseModel):
    name: str
    city: Optional[str] = ""
    status: str = "operational"
    is_head_office: bool = False
    description: Optional[str] = ""
    address: Optional[str] = ""
    image: Optional[str] = None
    image2: Optional[str] = None
    image3: Optional[str] = None
    gallery: Optional[List[str]] = []
    google_maps_url: Optional[str] = ""
    contact_phone: Optional[str] = ""
    whatsapp_number: Optional[str] = ""
    expected_opening: Optional[str] = ""
    order_index: int = 0
    is_offers_enabled: bool = True


class NotifyRequest(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    plaza_id: str
    name: str
    mobile: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class NotifyRequestCreate(BaseModel):
    plaza_id: str
    name: str
    mobile: str



class TimeSlot(BaseModel):
    open: str
    close: str


class Outlet(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    plaza_id: Optional[str] = None
    name: str
    mobile: str
    opening_time: str
    closing_time: str
    time_slots: Optional[List[TimeSlot]] = None
    logo: Optional[str] = None  # base64 or URL
    image2: Optional[str] = None  # extra gallery image #1 (optional)
    image3: Optional[str] = None  # extra gallery image #2 (optional)
    description: Optional[str] = ""
    address: Optional[str] = ""
    is_reservation_enabled: bool = False
    is_offers_enabled: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class OutletCreate(BaseModel):
    plaza_id: Optional[str] = None
    name: str
    mobile: str
    opening_time: str
    closing_time: str
    time_slots: Optional[List[TimeSlot]] = None
    logo: Optional[str] = None
    image2: Optional[str] = None
    image3: Optional[str] = None
    description: Optional[str] = ""
    address: Optional[str] = ""
    is_reservation_enabled: bool = False
    is_offers_enabled: bool = True


class MenuItem(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    outlet_id: str
    category: str
    name: str
    price: float
    image: Optional[str] = None
    description: Optional[str] = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class MenuItemCreate(BaseModel):
    outlet_id: str
    category: str
    name: str
    price: float
    image: Optional[str] = None
    description: Optional[str] = ""


class Offer(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str
    description: Optional[str] = ""
    image: Optional[str] = None
    outlet_id: Optional[str] = None  # null => applies to all outlets
    valid_until: Optional[str] = None
    is_active: bool = True
    is_offer_of_the_day: bool = False  # admin-picked "Offer of the Day" per outlet
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class OfferCreate(BaseModel):
    title: str
    description: Optional[str] = ""
    image: Optional[str] = None
    outlet_id: Optional[str] = None
    valid_until: Optional[str] = None
    is_active: bool = True
    is_offer_of_the_day: bool = False


class OfferClaim(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    mobile: str  # 10-digit Indian mobile (stored without prefix)
    dob: Optional[str] = ""  # YYYY-MM-DD
    anniversary: Optional[str] = ""  # YYYY-MM-DD
    plaza_id: Optional[str] = None
    outlet_id: Optional[str] = None
    offer_id: Optional[str] = None
    offer_title: Optional[str] = ""
    token: str
    valid_until: datetime  # end-of-day IST of creation
    offer_availed: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class OfferClaimCreate(BaseModel):
    name: str
    mobile: str
    dob: Optional[str] = ""
    anniversary: Optional[str] = ""
    plaza_id: Optional[str] = None
    outlet_id: Optional[str] = None
    offer_id: Optional[str] = None


class OfferClaimAvailedUpdate(BaseModel):
    offer_availed: bool


class Reservation(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    outlet_id: str
    name: str
    mobile: str
    date: str  # YYYY-MM-DD
    time: str  # HH:MM
    guests: int
    status: str = "pending"  # pending | confirmed | cancelled
    notes: Optional[str] = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class ReservationCreate(BaseModel):
    outlet_id: str
    name: str
    mobile: str
    date: str
    time: str
    guests: int
    notes: Optional[str] = ""


class ReservationStatusUpdate(BaseModel):
    status: str


class ClickEvent(BaseModel):
    outlet_id: str
    type: str  # "call" | "whatsapp"


class BulkMenuItem(BaseModel):
    outlet_id: Optional[str] = None
    outlet_name: Optional[str] = None  # fallback if outlet_id not known
    category: str
    name: str
    price: float
    image: Optional[str] = None
    description: Optional[str] = ""


class BulkMenuRequest(BaseModel):
    items: List[BulkMenuItem]
    replace_existing: bool = False  # if true, wipe existing menu of referenced outlets first


class Feedback(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    outlet_id: Optional[str] = None
    name: str
    mobile: Optional[str] = ""
    rating: int
    category: str  # Food | Service | Cleanliness | Ambience | Other
    message: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class FeedbackCreate(BaseModel):
    outlet_id: Optional[str] = None
    name: str
    mobile: Optional[str] = ""
    rating: int
    category: str
    message: str


class AboutContent(BaseModel):
    id: str = "about"
    title: str = "About Gallops Food Plaza"
    body: str = (
        "At Gallops Food Plaza, our journey began with a simple yet powerful idea – to "
        "redefine the highway dining experience. What started as a single destination has "
        "now grown into a network of dynamic food plazas across key highways, each offering "
        "a blend of quality, variety, hygiene, and hospitality.\n\n"
        "We've built more than just food courts. We've created culinary hubs that bring "
        "together leading brands, family-friendly spaces, and consistent cleanliness standards "
        "to offer travellers a break that's both memorable and refreshing.\n\n"
        "Every plaza reflects our passion for food, people, and excellence. From authentic "
        "Indian flavours to modern fusion fare, every bite is crafted with care, served with "
        "warmth, and backed by the highest food safety standards in the industry.\n\n"
        "Gallops isn't just where you stop to eat — it's where the journey becomes delicious."
    )
    hero_image: str = ""
    hero_banner: str = ""
    # Leadership section
    leadership: List[dict] = []
    # Why-choose bullets
    why_choose: List[str] = [
        "Fresh & Hygienic Food",
        "Multi-Cuisine Options",
        "EV Charging Stations",
        "Family-Friendly Spaces",
        "Convenient Highway Locations",
        "Clean Restrooms",
    ]
    franchise_title: str = "Franchise Enquiry"
    franchise_description: str = (
        "Partner with one of Gujarat's fastest-growing highway food plaza chains. "
        "Tap below to reach our franchise team."
    )
    franchise_phone: str = "+918779515804"
    franchise_email: str = ""
    updated_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


class AboutUpdate(BaseModel):
    title: Optional[str] = None
    body: Optional[str] = None
    hero_image: Optional[str] = None
    hero_banner: Optional[str] = None
    leadership: Optional[List[dict]] = None
    why_choose: Optional[List[str]] = None
    franchise_title: Optional[str] = None
    franchise_description: Optional[str] = None
    franchise_phone: Optional[str] = None
    franchise_email: Optional[str] = None


# ----------------- Auth Helpers -----------------
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def verify_password(plain: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))
    except Exception:
        return False


def create_access_token(user_id: str, email: str) -> str:
    payload = {
        "sub": user_id,
        "email": email,
        "exp": datetime.now(timezone.utc) + timedelta(days=7),
        "type": "access",
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)


async def get_current_admin(request: Request) -> dict:
    auth_header = request.headers.get("Authorization", "")
    token = auth_header[7:] if auth_header.startswith("Bearer ") else None
    # Fallback to ?token=<jwt> query param (used for file downloads / exports
    # where we can't easily set request headers).
    if not token:
        token = request.query_params.get("token")
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
        if not user or user.get("role") != "admin":
            raise HTTPException(status_code=403, detail="Admin access required")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")


# ----------------- Auth Routes -----------------
@api_router.post("/auth/login", response_model=LoginResponse)
async def login(body: LoginRequest):
    email = body.email.lower().strip()
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(body.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    token = create_access_token(user["id"], user["email"])
    user_out = {"id": user["id"], "email": user["email"], "name": user.get("name", ""), "role": user["role"]}
    return LoginResponse(token=token, user=user_out)


@api_router.get("/auth/me")
async def me(current: dict = Depends(get_current_admin)):
    return current


@api_router.post("/auth/change-password")
async def change_password(body: ChangePasswordRequest, current: dict = Depends(get_current_admin)):
    if len(body.new_password) < 6:
        raise HTTPException(status_code=400, detail="New password must be at least 6 characters")
    user = await db.users.find_one({"id": current["id"]})
    if not user or not verify_password(body.current_password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Current password is incorrect")
    await db.users.update_one(
        {"id": current["id"]},
        {"$set": {"password_hash": hash_password(body.new_password)}},
    )
    return {"success": True}


# ----------------- Plaza Routes -----------------
@api_router.get("/plazas", response_model=List[Plaza])
async def list_plazas(status: Optional[str] = Query(None)):
    q = {"status": status} if status else {}
    items = await db.plazas.find(q, {"_id": 0}).sort([("status", 1), ("order_index", 1), ("name", 1)]).to_list(200)
    return [Plaza(**p) for p in items]


# ----------------- Bootstrap (single-call hydrate for home) -----------------
# Fields stripped from outlets in the lite payload — these are heavy base64
# images that the home/plaza list view never renders. The plaza detail page
# then triggers a background full fetch to populate logos.
_LITE_OUTLET_OMIT = {"logo": 0, "image2": 0, "image3": 0, "description": 0}


# ---------------------------------------------------------------------------
# Outlet brand priority — flagship + key franchise brands always come first,
# everything else falls back to alphabetical. New plazas / new outlets pick
# up this ordering automatically. Add new brands to PRIORITY_BRANDS below.
# ---------------------------------------------------------------------------
import re as _re_brand  # noqa: E402

# Order matters — index = priority (lower = appears first).
PRIORITY_BRANDS: list[tuple[str, list[str]]] = [
    # (canonical_label, regex patterns matched against normalised name)
    ("Gallops Restaurant", [r"^gallops\s+restaurant$", r"^gallops$"]),
    # Match "Domino's", "Dominos", "Dominoz" (typo variant)
    ("Domino's",           [r"^domino[s'\u2019z]*\b"]),
    ("Subway",             [r"^subway\b"]),
    # Match "La Pino'z Pizza", "La Pinoz", "Lapinoz", "La Pino's", "Lapinoz Pizza"
    ("La Pino'z Pizza",    [r"^la\s*pino[\u2019'z]*\b", r"^lapino[\u2019'z]*\b"]),
    ("Lord Petrick",       [r"^lord\s+petrick\b"]),
    ("MMC",                [r"^mmc\b"]),
]
# Pre-compile for speed
_PRIORITY_REGEX = [
    (label, [_re_brand.compile(p, _re_brand.IGNORECASE) for p in patterns])
    for label, patterns in PRIORITY_BRANDS
]


def _outlet_priority(name: str) -> int:
    """Return the index (0..N-1) of this outlet's brand priority. Anything not
    matching any priority brand falls back to a high number so it sorts AFTER
    the priority brands (alphabetically among themselves)."""
    normalised = _re_brand.sub(r"\s+", " ", (name or "").strip().lower())
    for idx, (_label, patterns) in enumerate(_PRIORITY_REGEX):
        for pat in patterns:
            if pat.search(normalised):
                return idx
    return len(_PRIORITY_REGEX) + 100  # well above any brand index


def _sorted_outlets(outlets: list[dict]) -> list[dict]:
    """Sort outlets by (brand priority, name alphabetical, original created_at).

    `created_at` may be a `datetime` (modern Pydantic insert) OR an ISO `str`
    (early seed-script insert). Coerce to `str` so cross-type compares don't
    explode when ties happen on (priority, name).
    """
    return sorted(
        outlets,
        key=lambda o: (
            _outlet_priority(o.get("name", "")),
            (o.get("name") or "").strip().lower(),
            str(o.get("created_at") or ""),
        ),
    )


@api_router.get("/bootstrap")
async def bootstrap():
    """
    One-shot endpoint that returns plazas + LITE outlets + active offers in a
    single round-trip. Outlets are intentionally returned WITHOUT base64 logos
    or extra images (those are fetched lazily by the plaza/outlet detail
    screens). This keeps the cold-start payload small and makes navigation
    feel instant. Use pull-to-refresh to re-hydrate.
    """
    plazas_raw, outlets_raw, offers_raw = await asyncio.gather(
        db.plazas.find({}, {"_id": 0})
            .sort([("status", 1), ("order_index", 1), ("name", 1)])
            .to_list(200),
        # _LITE_OUTLET_OMIT excludes heavy fields server-side
        db.outlets.find({}, {"_id": 0, **_LITE_OUTLET_OMIT})
            .sort("created_at", 1).to_list(2000),
        db.offers.find({"is_active": True}, {"_id": 0}).to_list(500),
    )
    # Backfill required defaults on legacy seed docs that never wrote these
    # fields (notably the early `migrate_gallops_restaurant.py` outlets).
    for o in outlets_raw:
        o.setdefault("is_offers_enabled", True)
        o.setdefault("is_reservation_enabled", False)
        o.setdefault("time_slots", None)
    # Apply brand-priority sort so Gallops Restaurant / Domino's / Subway /
    # La Pino'z / Lord Petrick / MMC always come first inside each plaza.
    outlets_raw = _sorted_outlets(outlets_raw)
    return {
        "plazas": [Plaza(**p).dict() for p in plazas_raw],
        # outlets are lite — logos/extra images intentionally omitted
        "outlets": outlets_raw,
        "offers": [Offer(**o).dict() for o in offers_raw],
    }


@api_router.get("/plazas/{plaza_id}", response_model=Plaza)
async def get_plaza(plaza_id: str):
    p = await db.plazas.find_one({"id": plaza_id}, {"_id": 0})
    if not p:
        raise HTTPException(status_code=404, detail="Plaza not found")
    return Plaza(**p)


@api_router.post("/plazas", response_model=Plaza)
async def create_plaza(body: PlazaCreate, _: dict = Depends(get_current_admin)):
    plaza = Plaza(**body.dict())
    await db.plazas.insert_one(plaza.dict())
    return plaza


@api_router.put("/plazas/{plaza_id}", response_model=Plaza)
async def update_plaza(plaza_id: str, body: PlazaCreate, _: dict = Depends(get_current_admin)):
    res = await db.plazas.update_one({"id": plaza_id}, {"$set": body.dict()})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Plaza not found")
    p = await db.plazas.find_one({"id": plaza_id}, {"_id": 0})
    return Plaza(**p)


@api_router.delete("/plazas/{plaza_id}")
async def delete_plaza(plaza_id: str, _: dict = Depends(get_current_admin)):
    res = await db.plazas.delete_one({"id": plaza_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Plaza not found")
    # Cascade: unlink outlets from this plaza (do not delete them; safer)
    await db.outlets.update_many({"plaza_id": plaza_id}, {"$set": {"plaza_id": None}})
    return {"success": True}


# ----------------- Notify Requests (Upcoming Plazas) -----------------
@api_router.post("/notify-requests", response_model=NotifyRequest)
async def create_notify_request(body: NotifyRequestCreate):
    if not body.name.strip() or not body.mobile.strip():
        raise HTTPException(status_code=400, detail="Name and mobile are required")
    plaza = await db.plazas.find_one({"id": body.plaza_id})
    if not plaza:
        raise HTTPException(status_code=404, detail="Plaza not found")
    nr = NotifyRequest(**body.dict())
    await db.notify_requests.insert_one(nr.dict())
    return nr


@api_router.get("/admin/notify-requests", response_model=List[NotifyRequest])
async def list_notify_requests(
    plaza_id: Optional[str] = Query(None),
    _: dict = Depends(get_current_admin),
):
    q = {"plaza_id": plaza_id} if plaza_id else {}
    items = await db.notify_requests.find(q, {"_id": 0}).sort("created_at", -1).to_list(500)
    return [NotifyRequest(**n) for n in items]


@api_router.delete("/admin/notify-requests/{req_id}")
async def delete_notify_request(req_id: str, _: dict = Depends(get_current_admin)):
    res = await db.notify_requests.delete_one({"id": req_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Notify request not found")
    return {"success": True}


# ----------------- Outlet Routes -----------------
@api_router.get("/outlets")
async def list_outlets(
    plaza_id: Optional[str] = Query(None),
    lite: Optional[bool] = Query(False, description="If true, omit logo/image2/image3/description"),
):
    q = {"plaza_id": plaza_id} if plaza_id else {}
    projection = {"_id": 0}
    if lite:
        projection.update(_LITE_OUTLET_OMIT)
    items = await db.outlets.find(q, projection).sort("created_at", 1).to_list(1000)
    if lite:
        # Backfill required defaults on legacy seed docs that lack them so
        # the lite payload still satisfies the frontend's expected shape
        # without re-introducing the heavy fields we just stripped.
        for o in items:
            o.setdefault("is_offers_enabled", True)
            o.setdefault("is_reservation_enabled", False)
            o.setdefault("time_slots", None)
        # Apply brand-priority sort (Gallops Restaurant first, then Domino's,
        # Subway, La Pino'z, Lord Petrick, MMC, then alphabetical).
        items = _sorted_outlets(items)
        # Return raw dicts so the omitted fields stay omitted in the
        # serialised HTTP response (a `response_model=List[Outlet]`
        # decorator would otherwise re-inject them with default values).
        return items
    full = [Outlet(**o).dict() for o in items]
    return _sorted_outlets(full)


@api_router.get("/outlets/{outlet_id}", response_model=Outlet)
async def get_outlet(outlet_id: str):
    o = await db.outlets.find_one({"id": outlet_id}, {"_id": 0})
    if not o:
        raise HTTPException(status_code=404, detail="Outlet not found")
    return Outlet(**o)


@api_router.post("/outlets", response_model=Outlet)
async def create_outlet(body: OutletCreate, _: dict = Depends(get_current_admin)):
    outlet = Outlet(**body.dict())
    await db.outlets.insert_one(outlet.dict())
    return outlet


@api_router.put("/outlets/{outlet_id}", response_model=Outlet)
async def update_outlet(outlet_id: str, body: OutletCreate, _: dict = Depends(get_current_admin)):
    result = await db.outlets.update_one({"id": outlet_id}, {"$set": body.dict()})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Outlet not found")
    o = await db.outlets.find_one({"id": outlet_id}, {"_id": 0})
    return Outlet(**o)


@api_router.delete("/outlets/{outlet_id}")
async def delete_outlet(outlet_id: str, _: dict = Depends(get_current_admin)):
    result = await db.outlets.delete_one({"id": outlet_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Outlet not found")
    # cascade clean up menu & offers for this outlet
    await db.menu_items.delete_many({"outlet_id": outlet_id})
    await db.offers.delete_many({"outlet_id": outlet_id})
    return {"success": True}


# ----------------- Menu Routes -----------------
@api_router.get("/menu", response_model=List[MenuItem])
async def list_menu(outlet_id: Optional[str] = Query(None)):
    q = {"outlet_id": outlet_id} if outlet_id else {}
    items = await db.menu_items.find(q, {"_id": 0}).sort("category", 1).to_list(2000)
    # Filter out items whose (outlet_id, category) is marked hidden by admin
    hidden = await db.hidden_categories.find({}, {"_id": 0}).to_list(5000)
    hidden_set = {(h["outlet_id"], h["category"]) for h in hidden}
    items = [
        m for m in items
        if (m.get("outlet_id"), m.get("category")) not in hidden_set
    ]
    return [MenuItem(**m) for m in items]


@api_router.post("/menu", response_model=MenuItem)
async def create_menu_item(body: MenuItemCreate, _: dict = Depends(get_current_admin)):
    item = MenuItem(**body.dict())
    await db.menu_items.insert_one(item.dict())
    return item


@api_router.put("/menu/{item_id}", response_model=MenuItem)
async def update_menu_item(item_id: str, body: MenuItemCreate, _: dict = Depends(get_current_admin)):
    result = await db.menu_items.update_one({"id": item_id}, {"$set": body.dict()})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Menu item not found")
    m = await db.menu_items.find_one({"id": item_id}, {"_id": 0})
    return MenuItem(**m)


@api_router.delete("/menu/{item_id}")
async def delete_menu_item(item_id: str, _: dict = Depends(get_current_admin)):
    result = await db.menu_items.delete_one({"id": item_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Menu item not found")
    return {"success": True}


# ----------------- Categories Master (Admin) -----------------
# Hidden categories are stored in `hidden_categories`:
#   { outlet_id: str, category: str }
# A row's presence means the category is currently hidden from customers for that outlet.


@api_router.get("/admin/categories")
async def admin_list_categories(
    outlet_id: Optional[str] = Query(None),
    _: dict = Depends(get_current_admin),
):
    """
    Aggregate all categories across menu items grouped by outlet.
    Returns [{outlet_id, outlet_name, category, item_count, is_hidden}].
    If outlet_id is supplied, returns rows only for that outlet.
    """
    match: dict = {}
    if outlet_id:
        match["outlet_id"] = outlet_id
    pipeline = [
        {"$match": match} if match else {"$match": {}},
        {
            "$group": {
                "_id": {"outlet_id": "$outlet_id", "category": "$category"},
                "item_count": {"$sum": 1},
            }
        },
        {
            "$project": {
                "_id": 0,
                "outlet_id": "$_id.outlet_id",
                "category": "$_id.category",
                "item_count": 1,
            }
        },
        {"$sort": {"outlet_id": 1, "category": 1}},
    ]
    rows = await db.menu_items.aggregate(pipeline).to_list(5000)

    # Attach outlet names + hidden flag
    outlet_ids = list({r["outlet_id"] for r in rows if r.get("outlet_id")})
    outlets = await db.outlets.find({"id": {"$in": outlet_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(2000)
    name_map = {o["id"]: o["name"] for o in outlets}

    hidden = await db.hidden_categories.find({}, {"_id": 0}).to_list(5000)
    hidden_set = {(h["outlet_id"], h["category"]) for h in hidden}

    for r in rows:
        r["outlet_name"] = name_map.get(r.get("outlet_id"), "")
        r["is_hidden"] = (r.get("outlet_id"), r.get("category")) in hidden_set
    return rows


class CategoryVisibilityRequest(BaseModel):
    outlet_id: str
    category: str
    is_hidden: bool


@api_router.patch("/admin/categories/visibility")
async def admin_set_category_visibility(
    body: CategoryVisibilityRequest,
    _: dict = Depends(get_current_admin),
):
    key = {"outlet_id": body.outlet_id, "category": body.category}
    if body.is_hidden:
        await db.hidden_categories.update_one(key, {"$set": key}, upsert=True)
    else:
        await db.hidden_categories.delete_one(key)
    return {"success": True}


class CategoryRenameRequest(BaseModel):
    outlet_id: str
    old_category: str
    new_category: str


@api_router.patch("/admin/categories/rename")
async def admin_rename_category(
    body: CategoryRenameRequest,
    _: dict = Depends(get_current_admin),
):
    new_name = (body.new_category or "").strip()
    if not new_name:
        raise HTTPException(status_code=400, detail="New category name is required")
    result = await db.menu_items.update_many(
        {"outlet_id": body.outlet_id, "category": body.old_category},
        {"$set": {"category": new_name}},
    )
    # Also migrate the hidden flag if present
    await db.hidden_categories.update_many(
        {"outlet_id": body.outlet_id, "category": body.old_category},
        {"$set": {"category": new_name}},
    )
    return {"success": True, "items_updated": result.modified_count}


@api_router.delete("/admin/categories")
async def admin_delete_category(
    outlet_id: str = Query(...),
    category: str = Query(...),
    _: dict = Depends(get_current_admin),
):
    """Delete all menu items belonging to this (outlet, category) pair."""
    result = await db.menu_items.delete_many({"outlet_id": outlet_id, "category": category})
    await db.hidden_categories.delete_one({"outlet_id": outlet_id, "category": category})
    return {"success": True, "items_deleted": result.deleted_count}


# ----------------- Offers Routes -----------------
@api_router.get("/offers", response_model=List[Offer])
async def list_offers(outlet_id: Optional[str] = Query(None), only_active: bool = Query(True)):
    q: dict = {}
    if only_active:
        q["is_active"] = True
    if outlet_id:
        q["$or"] = [{"outlet_id": outlet_id}, {"outlet_id": None}]
    items = await db.offers.find(q, {"_id": 0}).sort("created_at", -1).to_list(500)
    return [Offer(**o) for o in items]


@api_router.post("/offers", response_model=Offer)
async def create_offer(body: OfferCreate, _: dict = Depends(get_current_admin)):
    offer = Offer(**body.dict())
    await db.offers.insert_one(offer.dict())
    return offer


@api_router.put("/offers/{offer_id}", response_model=Offer)
async def update_offer(offer_id: str, body: OfferCreate, _: dict = Depends(get_current_admin)):
    result = await db.offers.update_one({"id": offer_id}, {"$set": body.dict()})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Offer not found")
    o = await db.offers.find_one({"id": offer_id}, {"_id": 0})
    return Offer(**o)


@api_router.delete("/offers/{offer_id}")
async def delete_offer(offer_id: str, _: dict = Depends(get_current_admin)):
    result = await db.offers.delete_one({"id": offer_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Offer not found")
    return {"success": True}


# ----------------- Offer Claims Routes (Customer Login for Offer of the Day) -----------------
import secrets
import re
from io import BytesIO
from fastapi.responses import StreamingResponse


IST_OFFSET = timedelta(hours=5, minutes=30)


def _end_of_day_ist_utc(now_utc: Optional[datetime] = None) -> datetime:
    """Return the end-of-day (23:59:59) in IST, converted back to UTC."""
    if now_utc is None:
        now_utc = datetime.now(timezone.utc)
    ist = now_utc + IST_OFFSET
    eod_ist = datetime(ist.year, ist.month, ist.day, 23, 59, 59, tzinfo=timezone.utc)
    return eod_ist - IST_OFFSET


def _gen_token(length: int = 8) -> str:
    alphabet = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"  # no I/O/0/1 ambiguity
    return "".join(secrets.choice(alphabet) for _ in range(length))


def _normalise_mobile(raw: str) -> str:
    digits = re.sub(r"[^0-9]", "", raw or "")
    # strip country code 91 if 12 digits
    if len(digits) == 12 and digits.startswith("91"):
        digits = digits[2:]
    if len(digits) != 10:
        raise HTTPException(status_code=400, detail="Mobile must be a 10-digit Indian number")
    if digits[0] not in "6789":
        raise HTTPException(status_code=400, detail="Invalid Indian mobile number")
    return digits


def _normalise_user_date(raw: str) -> str:
    """
    Accept DD-MM-YYYY / DD/MM/YYYY / DD.MM.YYYY (or YYYY-MM-DD) and return
    an ISO date string (YYYY-MM-DD). Empty input returns "". Malformed
    input also returns "" (dates are optional in this context).
    """
    s = (raw or "").strip()
    if not s:
        return ""
    parts = re.split(r"[-/.]", s)
    if len(parts) != 3:
        return ""
    try:
        if len(parts[0]) == 4:
            y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
        else:
            d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
    except ValueError:
        return ""
    if not (1 <= d <= 31 and 1 <= m <= 12 and 1900 <= y <= 2100):
        return ""
    return f"{y:04d}-{m:02d}-{d:02d}"


def _pretty_user_date(iso: str) -> str:
    """Convert ISO YYYY-MM-DD to DD-MM-YYYY for user-facing display / exports."""
    s = (iso or "").strip()
    if not s:
        return ""
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", s)
    if not m:
        return s
    return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"


@api_router.post("/offer-claims")
async def create_offer_claim(body: OfferClaimCreate):
    """
    Customer-facing: captures Name, Mobile (+91), DOB (optional), Anniversary (optional).
    Returns the claim with a unique token + end-of-day validity + WhatsApp deep link.
    """
    name = (body.name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Name is required")
    mobile = _normalise_mobile(body.mobile)

    # --- One-mobile-per-day soft guard ---------------------------------------
    # If this mobile already claimed any offer today (valid_until still in the
    # future because claims expire at 23:59:59 IST), silently return the
    # EXISTING token instead of creating a duplicate. This prevents abuse
    # (multiple tokens per mobile) without hard-blocking the user — they simply
    # see the same token they already have.
    now_utc = datetime.now(timezone.utc)
    existing = await db.offer_claims.find_one(
        {"mobile": mobile, "valid_until": {"$gt": now_utc}},
        sort=[("created_at", -1)],
    )
    if existing:
        existing_outlet = None
        existing_plaza = None
        if existing.get("outlet_id"):
            existing_outlet = await db.outlets.find_one(
                {"id": existing["outlet_id"]}, {"_id": 0}
            )
        if existing.get("plaza_id"):
            existing_plaza = await db.plazas.find_one(
                {"id": existing["plaza_id"]}, {"_id": 0}
            )
        existing_offer = None
        if existing.get("offer_id"):
            existing_offer = await db.offers.find_one(
                {"id": existing["offer_id"]}, {"_id": 0}
            )
        vu = existing["valid_until"]
        if vu.tzinfo is None:
            vu = vu.replace(tzinfo=timezone.utc)
        valid_disp_existing = (vu + IST_OFFSET).strftime("%d-%b-%Y %H:%M IST")

        # Rebuild WhatsApp deep link for the existing claim so the customer
        # can re-send the same token to the plaza/outlet if they want.
        CUSTOMER_CARE_WHATSAPP = "919824006262"
        ex_outlet_name = (existing_outlet or {}).get("name", "")
        ex_outlet_wa = re.sub(r"[^0-9]", "", (existing_outlet or {}).get("mobile", "") or "")
        if ex_outlet_wa and len(ex_outlet_wa) == 10 and ex_outlet_wa[0] in "6789":
            ex_outlet_wa = "91" + ex_outlet_wa
        ex_plaza_wa_raw = (existing_plaza or {}).get("whatsapp_number", "") or (existing_plaza or {}).get("contact_phone", "") or ""
        ex_plaza_wa = re.sub(r"[^0-9]", "", ex_plaza_wa_raw)
        if ex_plaza_wa and len(ex_plaza_wa) == 10 and ex_plaza_wa[0] in "6789":
            ex_plaza_wa = "91" + ex_plaza_wa
        ex_plaza_name = (existing_plaza or {}).get("name", "")
        if ex_outlet_wa:
            ex_wa_recipient = ex_outlet_wa
            ex_routed_to = ex_outlet_name or ex_plaza_name
        elif ex_plaza_wa:
            ex_wa_recipient = ex_plaza_wa
            ex_routed_to = ex_plaza_name
        else:
            ex_wa_recipient = CUSTOMER_CARE_WHATSAPP
            ex_routed_to = "Gallops Customer Care"
        ex_venue = ex_outlet_name or ex_plaza_name or "Gallops Food Plaza"
        ex_msg_lines = [
            f"Hello, I would like to claim the Offer of the Day at Gallops Food Plaza - {ex_venue}.",
            f"Token: {existing['token']}",
            f"Mobile: +91{mobile}",
            f"Valid Till: {valid_disp_existing}",
            "Message sent from Gallops Food Plaza Mobile App",
        ]
        ex_wa_text = "\n".join(ex_msg_lines)
        ex_wa_link = f"https://wa.me/{ex_wa_recipient}?text={wa_text_encode(ex_wa_text)}"

        return {
            "id": existing["id"],
            "token": existing["token"],
            "valid_until": vu.isoformat(),
            "valid_until_display": valid_disp_existing,
            "offer": existing_offer,
            "outlet_name": ex_outlet_name,
            "plaza_name": ex_plaza_name,
            "routed_to": ex_routed_to,
            "routed_tier": "existing",
            "whatsapp_link": ex_wa_link,
            "whatsapp_message": ex_wa_text,
            "mobile": mobile,
            "already_claimed": True,
            "message": "You have already claimed an offer today. Here is your existing token — it is valid until end of day.",
        }

    # Resolve the offer-of-the-day & outlet context
    outlet = None
    offer = None
    if body.outlet_id:
        outlet = await db.outlets.find_one({"id": body.outlet_id}, {"_id": 0})
    # Prefer explicit offer_id sent from the client; otherwise try to auto-resolve OTD for outlet
    if body.offer_id:
        offer = await db.offers.find_one({"id": body.offer_id, "is_active": True}, {"_id": 0})
    elif body.outlet_id:
        offer = await db.offers.find_one(
            {
                "outlet_id": body.outlet_id,
                "is_active": True,
                "is_offer_of_the_day": True,
            },
            {"_id": 0},
        )

    plaza_id = body.plaza_id or (outlet.get("plaza_id") if outlet else None)
    plaza = await db.plazas.find_one({"id": plaza_id}, {"_id": 0}) if plaza_id else None

    token = _gen_token()
    # ensure uniqueness (extremely unlikely collision, but just in case)
    for _ in range(5):
        if not await db.offer_claims.find_one({"token": token}):
            break
        token = _gen_token()

    valid_until_utc = _end_of_day_ist_utc()

    claim = OfferClaim(
        name=name,
        mobile=mobile,
        dob=_normalise_user_date(body.dob or ""),
        anniversary=_normalise_user_date(body.anniversary or ""),
        plaza_id=plaza_id,
        outlet_id=body.outlet_id,
        offer_id=(offer or {}).get("id"),
        offer_title=(offer or {}).get("title", ""),
        token=token,
        valid_until=valid_until_utc,
    )
    await db.offer_claims.insert_one(claim.dict())

    # Build WhatsApp deep-link with routing priority:
    #   1. Outlet mobile (if claim is for a specific outlet)
    #   2. Plaza's whatsapp_number (if configured in Admin)
    #   3. Customer Care number (last-resort fallback — see CUSTOMER_CARE_WHATSAPP)
    CUSTOMER_CARE_WHATSAPP = "919824006262"  # +91 98240 06262 (Gallops Customer Care)

    outlet_name = (outlet or {}).get("name", "")
    outlet_wa = re.sub(r"[^0-9]", "", (outlet or {}).get("mobile", "") or "")
    if outlet_wa and len(outlet_wa) == 10 and outlet_wa[0] in "6789":
        outlet_wa = "91" + outlet_wa
    # Prefer explicit plaza whatsapp_number; if empty, silently fall back to
    # the plaza's contact_phone (most Indian plazas share the same number
    # for calls + WhatsApp). This means just configuring Contact Phone on a
    # plaza is already enough to route offer messages to that plaza.
    plaza_wa_raw = (plaza or {}).get("whatsapp_number", "") or (plaza or {}).get("contact_phone", "") or ""
    plaza_wa = re.sub(r"[^0-9]", "", plaza_wa_raw)
    if plaza_wa and len(plaza_wa) == 10 and plaza_wa[0] in "6789":
        plaza_wa = "91" + plaza_wa  # prefix country code if the admin stored a bare 10-digit
    plaza_name = (plaza or {}).get("name", "")

    # Pick the recipient in priority order
    if outlet_wa:
        wa_recipient = outlet_wa
        routed_to = outlet_name or plaza_name
        routed_tier = "outlet"
    elif plaza_wa:
        wa_recipient = plaza_wa
        routed_to = plaza_name
        routed_tier = "plaza"
    else:
        wa_recipient = CUSTOMER_CARE_WHATSAPP
        routed_to = "Gallops Customer Care"
        routed_tier = "customer_care"

    # Format validity in IST for humans
    eod_ist = valid_until_utc + IST_OFFSET
    valid_disp = eod_ist.strftime("%d-%b-%Y %H:%M IST")
    # The outlet/plaza name shown in the message — prefer the specific outlet
    # when there is one, else plaza.
    venue_name = outlet_name or plaza_name or "Gallops Food Plaza"
    msg_lines = [
        f"Hello, I would like to claim the Offer of the Day at Gallops Food Plaza - {venue_name}.",
        f"Token: {token}",
        f"Mobile: +91{mobile}",
        f"Valid Till: {valid_disp}",
        "Message sent from Gallops Food Plaza Mobile App",
    ]
    wa_text = "\n".join(msg_lines)
    wa_link = f"https://wa.me/{wa_recipient}?text={wa_text_encode(wa_text)}"

    return {
        "id": claim.id,
        "token": token,
        "valid_until": valid_until_utc.isoformat(),
        "valid_until_display": valid_disp,
        "offer": offer,  # may be None if no OTD configured
        "outlet_name": outlet_name,
        "plaza_name": plaza_name,
        "routed_to": routed_to,          # human-readable destination
        "routed_tier": routed_tier,      # 'outlet' | 'plaza' | 'customer_care'
        "whatsapp_link": wa_link,
        "whatsapp_message": wa_text,
        "mobile": mobile,
    }


def wa_text_encode(s: str) -> str:
    # WhatsApp accepts standard URI encoding of the text
    from urllib.parse import quote
    return quote(s)


@api_router.get("/admin/offer-claims")
async def list_offer_claims(
    date_from: Optional[str] = Query(None),  # YYYY-MM-DD (IST)
    date_to: Optional[str] = Query(None),    # YYYY-MM-DD (IST)
    plaza_id: Optional[str] = Query(None),
    outlet_id: Optional[str] = Query(None),
    mobile: Optional[str] = Query(None),
    availed: Optional[str] = Query(None),  # "yes" | "no"
    limit: int = Query(500, ge=1, le=5000),
    _: dict = Depends(get_current_admin),
):
    q: dict = {}
    if plaza_id:
        q["plaza_id"] = plaza_id
    if outlet_id:
        q["outlet_id"] = outlet_id
    if mobile:
        digits = re.sub(r"[^0-9]", "", mobile)
        if digits:
            q["mobile"] = {"$regex": digits}
    if availed == "yes":
        q["offer_availed"] = True
    elif availed == "no":
        q["offer_availed"] = False
    if date_from or date_to:
        created: dict = {}
        if date_from:
            try:
                d0 = datetime.strptime(date_from, "%Y-%m-%d")
                start_ist = datetime(d0.year, d0.month, d0.day, 0, 0, 0, tzinfo=timezone.utc)
                created["$gte"] = start_ist - IST_OFFSET
            except ValueError:
                pass
        if date_to:
            try:
                d1 = datetime.strptime(date_to, "%Y-%m-%d")
                end_ist = datetime(d1.year, d1.month, d1.day, 23, 59, 59, tzinfo=timezone.utc)
                created["$lte"] = end_ist - IST_OFFSET
            except ValueError:
                pass
        if created:
            q["created_at"] = created
    items = (
        await db.offer_claims.find(q, {"_id": 0}).sort("created_at", -1).to_list(limit)
    )
    # Enrich with outlet/plaza names
    outlet_map = {o["id"]: o for o in await db.outlets.find({}, {"_id": 0}).to_list(5000)}
    plaza_map = {p["id"]: p for p in await db.plazas.find({}, {"_id": 0}).to_list(200)}
    for it in items:
        if isinstance(it.get("created_at"), datetime):
            it["created_at"] = it["created_at"].isoformat()
        if isinstance(it.get("valid_until"), datetime):
            it["valid_until"] = it["valid_until"].isoformat()
        o = outlet_map.get(it.get("outlet_id") or "")
        p = plaza_map.get(it.get("plaza_id") or (o or {}).get("plaza_id") or "")
        it["outlet_name"] = (o or {}).get("name", "")
        it["plaza_name"] = (p or {}).get("name", "")
    return items


@api_router.patch("/admin/offer-claims/{claim_id}")
async def update_offer_claim_availed(
    claim_id: str,
    body: OfferClaimAvailedUpdate,
    _: dict = Depends(get_current_admin),
):
    res = await db.offer_claims.update_one(
        {"id": claim_id}, {"$set": {"offer_availed": bool(body.offer_availed)}}
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Claim not found")
    return {"success": True}


@api_router.delete("/admin/offer-claims/{claim_id}")
async def delete_offer_claim(claim_id: str, _: dict = Depends(get_current_admin)):
    res = await db.offer_claims.delete_one({"id": claim_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Claim not found")
    return {"success": True}


@api_router.get("/admin/offer-claims/export")
async def export_offer_claims(
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    plaza_id: Optional[str] = Query(None),
    outlet_id: Optional[str] = Query(None),
    mobile: Optional[str] = Query(None),
    availed: Optional[str] = Query(None),
    token: Optional[str] = Query(None),  # allow token-as-query for file downloads
    _: dict = Depends(get_current_admin),
):
    items = await list_offer_claims(
        date_from=date_from,
        date_to=date_to,
        plaza_id=plaza_id,
        outlet_id=outlet_id,
        mobile=mobile,
        availed=availed,
        limit=5000,
        _={"ok": True},  # type: ignore
    )

    # Build xlsx
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Offer Claims"

    headers = [
        "Date of Entry (IST)",
        "Name",
        "Mobile Number",
        "DOB",
        "Anniversary",
        "Plaza",
        "Outlet",
        "Offer",
        "Token",
        "Valid Till (IST)",
        "Offer Availed",
    ]
    ws.append(headers)
    head_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center")

    def ist_str(iso: str) -> str:
        try:
            dt = datetime.fromisoformat(iso.replace("Z", "+00:00"))
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return (dt + IST_OFFSET).strftime("%d-%b-%Y %H:%M")
        except Exception:
            return iso or ""

    for it in items:
        ws.append([
            ist_str(it.get("created_at", "")),
            it.get("name", ""),
            "+91" + (it.get("mobile", "") or ""),
            _pretty_user_date(it.get("dob", "") or ""),
            _pretty_user_date(it.get("anniversary", "") or ""),
            it.get("plaza_name", ""),
            it.get("outlet_name", ""),
            it.get("offer_title", ""),
            it.get("token", ""),
            ist_str(it.get("valid_until", "")),
            "Yes" if it.get("offer_availed") else "No",
        ])

    # Auto width
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(header)
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        for cell in ws[col_letter]:
            v = str(cell.value or "")
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"gallops-offer-claims-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ----------------- Reservation Routes -----------------
@api_router.post("/reservations", response_model=Reservation)
async def create_reservation(body: ReservationCreate):
    outlet = await db.outlets.find_one({"id": body.outlet_id})
    if not outlet:
        raise HTTPException(status_code=404, detail="Outlet not found")
    if not outlet.get("is_reservation_enabled"):
        raise HTTPException(status_code=400, detail="Reservations not available for this outlet")
    if body.guests < 1 or body.guests > 50:
        raise HTTPException(status_code=400, detail="Guests must be between 1 and 50")
    reservation = Reservation(**body.dict())
    await db.reservations.insert_one(reservation.dict())
    return reservation


@api_router.get("/reservations", response_model=List[Reservation])
async def list_reservations(
    outlet_id: Optional[str] = Query(None),
    date: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    _: dict = Depends(get_current_admin),
):
    q: dict = {}
    if outlet_id:
        q["outlet_id"] = outlet_id
    if date:
        q["date"] = date
    if status:
        q["status"] = status
    items = await db.reservations.find(q, {"_id": 0}).sort("created_at", -1).to_list(2000)
    return [Reservation(**r) for r in items]


@api_router.put("/reservations/{res_id}/status", response_model=Reservation)
async def update_reservation_status(
    res_id: str, body: ReservationStatusUpdate, _: dict = Depends(get_current_admin)
):
    if body.status not in {"pending", "confirmed", "cancelled"}:
        raise HTTPException(status_code=400, detail="Invalid status")
    result = await db.reservations.update_one({"id": res_id}, {"$set": {"status": body.status}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Reservation not found")
    r = await db.reservations.find_one({"id": res_id}, {"_id": 0})
    return Reservation(**r)


@api_router.delete("/reservations/{res_id}")
async def delete_reservation(res_id: str, _: dict = Depends(get_current_admin)):
    result = await db.reservations.delete_one({"id": res_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Reservation not found")
    return {"success": True}


# ----------------- Click Tracking -----------------
@api_router.post("/track/click")
async def track_click(body: ClickEvent):
    if body.type not in {"call", "whatsapp"}:
        raise HTTPException(status_code=400, detail="Invalid click type")
    await db.click_events.insert_one({
        "id": str(uuid.uuid4()),
        "outlet_id": body.outlet_id,
        "type": body.type,
        "created_at": datetime.now(timezone.utc),
    })
    return {"ok": True}


# ----------------- Admin Analytics -----------------
@api_router.get("/admin/analytics")
async def analytics(_: dict = Depends(get_current_admin)):
    now = datetime.now(timezone.utc)
    start_today = datetime(now.year, now.month, now.day, tzinfo=timezone.utc)
    start_week = start_today - timedelta(days=7)
    start_month = start_today - timedelta(days=30)

    totals = {
        "outlets": await db.outlets.count_documents({}),
        "menu_items": await db.menu_items.count_documents({}),
        "offers": await db.offers.count_documents({"is_active": True}),
        "reservations": await db.reservations.count_documents({}),
    }

    res_today = await db.reservations.count_documents({"created_at": {"$gte": start_today}})
    res_week = await db.reservations.count_documents({"created_at": {"$gte": start_week}})
    res_month = await db.reservations.count_documents({"created_at": {"$gte": start_month}})

    outlet_names = {
        o["id"]: o["name"]
        for o in await db.outlets.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(1000)
    }

    pipeline_res_outlet = [
        {"$group": {"_id": "$outlet_id", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 10},
    ]
    raw = await db.reservations.aggregate(pipeline_res_outlet).to_list(10)
    reservations_by_outlet = [
        {"outlet_id": r["_id"], "name": outlet_names.get(r["_id"], "Unknown"), "count": r["count"]}
        for r in raw
    ]

    by_status_raw = await db.reservations.aggregate(
        [{"$group": {"_id": "$status", "count": {"$sum": 1}}}]
    ).to_list(10)
    reservations_by_status = {r["_id"]: r["count"] for r in by_status_raw}

    clicks_raw = await db.click_events.aggregate(
        [{"$group": {"_id": {"outlet_id": "$outlet_id", "type": "$type"}, "count": {"$sum": 1}}}]
    ).to_list(2000)
    clicks_map: dict = {}
    for c in clicks_raw:
        oid = c["_id"]["outlet_id"]
        entry = clicks_map.setdefault(
            oid, {"outlet_id": oid, "name": outlet_names.get(oid, "Unknown"), "calls": 0, "whatsapps": 0}
        )
        if c["_id"]["type"] == "call":
            entry["calls"] = c["count"]
        else:
            entry["whatsapps"] = c["count"]
    clicks_by_outlet = sorted(
        clicks_map.values(), key=lambda x: x["calls"] + x["whatsapps"], reverse=True
    )[:10]
    total_calls = sum(c["calls"] for c in clicks_map.values())
    total_whatsapps = sum(c["whatsapps"] for c in clicks_map.values())

    recent = await db.reservations.find({}, {"_id": 0}).sort("created_at", -1).to_list(5)
    for r in recent:
        r["outlet_name"] = outlet_names.get(r.get("outlet_id", ""), "Unknown")
        if isinstance(r.get("created_at"), datetime):
            r["created_at"] = r["created_at"].isoformat()

    # Feedback stats
    feedback_count = await db.feedbacks.count_documents({})
    fb_agg = await db.feedbacks.aggregate([
        {"$group": {"_id": None, "avg": {"$avg": "$rating"}}}
    ]).to_list(1)
    avg_rating = round(fb_agg[0]["avg"], 2) if fb_agg else 0
    fb_by_outlet_raw = await db.feedbacks.aggregate([
        {"$match": {"outlet_id": {"$ne": None}}},
        {"$group": {"_id": "$outlet_id", "count": {"$sum": 1}, "avg": {"$avg": "$rating"}}},
        {"$sort": {"avg": -1}},
        {"$limit": 10},
    ]).to_list(10)
    feedback_by_outlet = [
        {
            "outlet_id": r["_id"],
            "name": outlet_names.get(r["_id"], "Unknown"),
            "count": r["count"],
            "avg_rating": round(r["avg"], 2),
        }
        for r in fb_by_outlet_raw
    ]

    return {
        "totals": totals,
        "reservations_by_period": {"today": res_today, "week": res_week, "month": res_month},
        "reservations_by_outlet": reservations_by_outlet,
        "reservations_by_status": reservations_by_status,
        "clicks_total": {"calls": total_calls, "whatsapps": total_whatsapps},
        "clicks_by_outlet": clicks_by_outlet,
        "recent_reservations": recent,
        "feedback_count": feedback_count,
        "feedback_avg_rating": avg_rating,
        "feedback_by_outlet": feedback_by_outlet,
    }


# ----------------- Bulk Menu Import -----------------
@api_router.post("/menu/bulk")
async def bulk_menu_import(body: BulkMenuRequest, _: dict = Depends(get_current_admin)):
    outlets = await db.outlets.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(1000)
    name_to_id = {o["name"].strip().lower(): o["id"] for o in outlets}
    valid_ids = {o["id"] for o in outlets}

    docs = []
    errors = []
    outlet_ids_touched = set()
    for i, item in enumerate(body.items):
        outlet_id = item.outlet_id
        if not outlet_id and item.outlet_name:
            outlet_id = name_to_id.get(item.outlet_name.strip().lower())
        if not outlet_id or outlet_id not in valid_ids:
            errors.append({"row": i + 1, "error": f"Unknown outlet: {item.outlet_name or item.outlet_id}"})
            continue
        if not item.name or not item.category or item.price is None:
            errors.append({"row": i + 1, "error": "Missing name/category/price"})
            continue
        outlet_ids_touched.add(outlet_id)
        docs.append({
            "id": str(uuid.uuid4()),
            "outlet_id": outlet_id,
            "category": item.category.strip(),
            "name": item.name.strip(),
            "price": float(item.price),
            "image": item.image or None,
            "description": item.description or "",
            "created_at": datetime.now(timezone.utc),
        })

    if body.replace_existing and outlet_ids_touched:
        await db.menu_items.delete_many({"outlet_id": {"$in": list(outlet_ids_touched)}})
    if docs:
        await db.menu_items.insert_many(docs)

    return {
        "inserted": len(docs),
        "errors": errors,
        "outlets_affected": len(outlet_ids_touched),
    }


# ----------------- Feedback -----------------
@api_router.post("/feedback", response_model=Feedback)
async def create_feedback(body: FeedbackCreate):
    if not 1 <= body.rating <= 5:
        raise HTTPException(status_code=400, detail="Rating must be between 1 and 5")
    if not body.name.strip() or not body.message.strip():
        raise HTTPException(status_code=400, detail="Name and message are required")
    if body.outlet_id:
        exists = await db.outlets.find_one({"id": body.outlet_id})
        if not exists:
            raise HTTPException(status_code=404, detail="Outlet not found")
    fb = Feedback(**body.dict())
    await db.feedbacks.insert_one(fb.dict())
    return fb


@api_router.get("/admin/feedback", response_model=List[Feedback])
async def list_feedback(
    outlet_id: Optional[str] = Query(None),
    min_rating: Optional[int] = Query(None),
    _: dict = Depends(get_current_admin),
):
    q: dict = {}
    if outlet_id:
        q["outlet_id"] = outlet_id
    if min_rating:
        q["rating"] = {"$gte": min_rating}
    items = await db.feedbacks.find(q, {"_id": 0}).sort("created_at", -1).to_list(500)
    return [Feedback(**f) for f in items]


@api_router.delete("/admin/feedback/{fb_id}")
async def delete_feedback(fb_id: str, _: dict = Depends(get_current_admin)):
    result = await db.feedbacks.delete_one({"id": fb_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Feedback not found")
    return {"success": True}


# ----------------- About / Franchise content -----------------
@api_router.get("/about", response_model=AboutContent)
async def get_about():
    doc = await db.settings.find_one({"id": "about"}, {"_id": 0})
    if not doc:
        # return defaults without writing (seeded on startup)
        return AboutContent()
    return AboutContent(**doc)


@api_router.put("/admin/about", response_model=AboutContent)
async def update_about(body: AboutUpdate, _: dict = Depends(get_current_admin)):
    existing = await db.settings.find_one({"id": "about"}, {"_id": 0}) or {}
    merged = {**AboutContent().dict(), **existing}
    updates = {k: v for k, v in body.dict().items() if v is not None}
    merged.update(updates)
    merged["id"] = "about"
    merged["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.settings.update_one({"id": "about"}, {"$set": merged}, upsert=True)
    return AboutContent(**merged)


# ----------------- Root -----------------
@api_router.get("/")
async def root():
    return {"message": "Gallops Food Plaza API"}


app.include_router(api_router)


# ----------------- Public Legal Pages (Play Store / App Store compliance) ----
# These routes are intentionally NOT under `/api/` so the URLs are clean and
# directly shareable as Privacy Policy / Terms URLs in the store listings.
from fastapi.responses import HTMLResponse  # noqa: E402

_LEGAL_DIR = os.path.join(os.path.dirname(__file__), "..", "store_assets")


def _render_legal_html(title: str, md_path: str) -> str:
    try:
        with open(md_path, "r", encoding="utf-8") as f:
            md = f.read()
    except FileNotFoundError:
        md = "*Document not yet uploaded.*"

    # Tiny markdown -> HTML converter (no extra dependency). Handles headings,
    # bold, italics, lists and paragraphs — sufficient for our legal copy.
    import re as _re
    html_lines: list[str] = []
    in_list = False
    for raw in md.split("\n"):
        line = raw.rstrip()
        if not line.strip():
            if in_list:
                html_lines.append("</ul>")
                in_list = False
            html_lines.append("")
            continue
        m_h = _re.match(r"^(#{1,6})\s+(.*)$", line)
        if m_h:
            if in_list:
                html_lines.append("</ul>")
                in_list = False
            level = len(m_h.group(1))
            html_lines.append(f"<h{level}>{m_h.group(2)}</h{level}>")
            continue
        if line.startswith("- ") or line.startswith("* "):
            if not in_list:
                html_lines.append("<ul>")
                in_list = True
            item = line[2:]
            item = _re.sub(r"\*\*(.+?)\*\*", r"<strong>\1</strong>", item)
            item = _re.sub(r"\*(.+?)\*", r"<em>\1</em>", item)
            html_lines.append(f"<li>{item}</li>")
            continue
        if line.strip() == "---":
            if in_list:
                html_lines.append("</ul>")
                in_list = False
            html_lines.append("<hr/>")
            continue
        # Paragraph
        if in_list:
            html_lines.append("</ul>")
            in_list = False
        para = _re.sub(r"\*\*(.+?)\*\*", r"<strong>\1</strong>", line)
        para = _re.sub(r"\*(.+?)\*", r"<em>\1</em>", para)
        html_lines.append(f"<p>{para}</p>")
    if in_list:
        html_lines.append("</ul>")

    body_html = "\n".join(html_lines)
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8" />
<meta name="viewport" content="width=device-width,initial-scale=1" />
<title>{title} — Gallops Food Plaza</title>
<style>
  :root {{ color-scheme: light; }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Helvetica, Arial, sans-serif;
         max-width: 760px; margin: 0 auto; padding: 32px 20px 80px; line-height: 1.65;
         color: #1f2937; background: #fafafa; }}
  h1 {{ font-size: 28px; margin-bottom: 6px; color: #1e1b4b; }}
  h2 {{ font-size: 20px; margin-top: 30px; color: #1e1b4b; }}
  h3 {{ font-size: 16px; margin-top: 22px; color: #1e1b4b; }}
  hr {{ border: 0; border-top: 1px solid #e5e7eb; margin: 28px 0; }}
  a  {{ color: #d63419; }}
  ul {{ padding-left: 20px; }}
  li {{ margin: 4px 0; }}
  p strong {{ color: #1e1b4b; }}
  .badge {{ display: inline-block; padding: 4px 10px; background: #eef2ff; color: #1e1b4b;
            border-radius: 999px; font-size: 12px; font-weight: 600; margin-bottom: 18px; }}
  footer {{ margin-top: 40px; font-size: 12px; color: #6b7280; text-align: center; }}
</style>
</head>
<body>
<span class="badge">GALLOPS FOOD PLAZA</span>
{body_html}
<footer>© Gallops Food Plaza. For questions, contact +91 77024 21761 or Fedra@gallopsfoodplaza.in.</footer>
</body>
</html>
"""


@app.get("/api/privacy", response_class=HTMLResponse, include_in_schema=False)
async def privacy_policy_page():
    return _render_legal_html(
        "Privacy Policy",
        _os.path.join(_LEGAL_DIR, "PRIVACY_POLICY.md"),
    )


@app.get("/api/terms", response_class=HTMLResponse, include_in_schema=False)
async def terms_of_service_page():
    return _render_legal_html(
        "Terms of Service",
        _os.path.join(_LEGAL_DIR, "TERMS_OF_SERVICE.md"),
    )

# Mount static assets (logos, images, etc.) — accessible at /api/static/<file>
import os as _os
_static_dir = _os.path.join(_os.path.dirname(__file__), "static")
if _os.path.isdir(_static_dir):
    app.mount("/api/static", StaticFiles(directory=_static_dir), name="static")

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)
# GZip every response > 1 KB. Most of our payloads are JSON with base64
# image strings which compress 2-3x — this dramatically speeds up plaza/outlet
# list fetches on slow networks.
app.add_middleware(GZipMiddleware, minimum_size=1024)


# ----------------- Seed Data -----------------
SEED_OUTLETS = [
    {
        "name": "Gallops Restaurant",
        "mobile": "+919876543210",
        "opening_time": "11:00",
        "closing_time": "23:30",
        "description": "Our signature multi-cuisine fine-dine restaurant. Reservations recommended.",
        "address": "Gallops Food Plaza, Anand",
        "logo": "https://images.unsplash.com/photo-1517248135467-4c7edcad34c4?w=800&q=80",
        "is_reservation_enabled": True,
    },
    {
        "name": "Gallops Cafe",
        "mobile": "+919876543211",
        "opening_time": "08:00",
        "closing_time": "22:00",
        "description": "Hand-brewed coffee, teas and freshly baked pastries.",
        "address": "Gallops Food Plaza, Anand",
        "logo": "https://images.unsplash.com/photo-1763683944193-c4635a3d3999?w=800&q=80",
        "is_reservation_enabled": False,
    },
    {
        "name": "Gallops Fast Food",
        "mobile": "+919876543212",
        "opening_time": "10:00",
        "closing_time": "23:00",
        "description": "Burgers, fries, pizzas and all-day munchies.",
        "address": "Gallops Food Plaza, Anand",
        "logo": "https://images.unsplash.com/photo-1632898657999-ae6920976661?w=800&q=80",
        "is_reservation_enabled": False,
    },
    {
        "name": "Gallops Juice Bar",
        "mobile": "+919876543213",
        "opening_time": "09:00",
        "closing_time": "22:00",
        "description": "Fresh fruit juices, smoothies and health shots.",
        "address": "Gallops Food Plaza, Anand",
        "logo": "https://images.unsplash.com/photo-1622597467836-f3e6f9f2bbbd?w=800&q=80",
        "is_reservation_enabled": False,
    },
    {
        "name": "Gallops Ice Cream",
        "mobile": "+919876543214",
        "opening_time": "11:00",
        "closing_time": "23:30",
        "description": "Artisanal ice creams, sundaes and desserts.",
        "address": "Gallops Food Plaza, Anand",
        "logo": "https://images.unsplash.com/photo-1501443762994-82bd5dace89a?w=800&q=80",
        "is_reservation_enabled": False,
    },
]


SEED_MENU_BY_OUTLET = {
    "Gallops Restaurant": [
        ("Starters", "Paneer Tikka", 320, "https://images.unsplash.com/photo-1599487488170-d11ec9c172f0?w=600&q=80"),
        ("Starters", "Crispy Corn", 260, "https://images.unsplash.com/photo-1626074353765-517a681e40be?w=600&q=80"),
        ("Main Course", "Butter Chicken", 480, "https://images.unsplash.com/photo-1603894584373-5ac82b2ae398?w=600&q=80"),
        ("Main Course", "Dal Makhani", 280, "https://images.unsplash.com/photo-1546833999-b9f581a1996d?w=600&q=80"),
        ("Breads", "Butter Naan", 60, "https://images.unsplash.com/photo-1626500155770-694f08a34e62?w=600&q=80"),
        ("Desserts", "Gulab Jamun", 120, "https://images.unsplash.com/photo-1601050690597-df0568f70950?w=600&q=80"),
    ],
    "Gallops Cafe": [
        ("Coffee", "Cappuccino", 160, "https://images.unsplash.com/photo-1509042239860-f550ce710b93?w=600&q=80"),
        ("Coffee", "Latte", 180, "https://images.unsplash.com/photo-1541167760496-1628856ab772?w=600&q=80"),
        ("Tea", "Masala Chai", 80, "https://images.unsplash.com/photo-1571934811356-5cc061b6821f?w=600&q=80"),
        ("Bakery", "Croissant", 120, "https://images.unsplash.com/photo-1555507036-ab1f4038808a?w=600&q=80"),
        ("Bakery", "Chocolate Muffin", 140, "https://images.unsplash.com/photo-1604882406195-c4f0c3bc1c7a?w=600&q=80"),
    ],
    "Gallops Fast Food": [
        ("Burgers", "Classic Veg Burger", 140, "https://images.unsplash.com/photo-1550547660-d9450f859349?w=600&q=80"),
        ("Burgers", "Cheese Burst Burger", 180, "https://images.unsplash.com/photo-1568901346375-23c9450c58cd?w=600&q=80"),
        ("Pizza", "Margherita Pizza", 260, "https://images.unsplash.com/photo-1574071318508-1cdbab80d002?w=600&q=80"),
        ("Pizza", "Farmhouse Pizza", 320, "https://images.unsplash.com/photo-1565299624946-b28f40a0ae38?w=600&q=80"),
        ("Sides", "French Fries", 120, "https://images.unsplash.com/photo-1576107232684-1279f390859f?w=600&q=80"),
    ],
    "Gallops Juice Bar": [
        ("Juices", "Orange Juice", 100, "https://images.unsplash.com/photo-1600271886742-f049cd451bba?w=600&q=80"),
        ("Juices", "Watermelon Juice", 90, "https://images.unsplash.com/photo-1622484211970-63dc3d6adc0d?w=600&q=80"),
        ("Smoothies", "Mango Smoothie", 150, "https://images.unsplash.com/photo-1623065422902-30a2d299bbe4?w=600&q=80"),
        ("Smoothies", "Berry Blast", 160, "https://images.unsplash.com/photo-1570696516188-ade861b84a49?w=600&q=80"),
    ],
    "Gallops Ice Cream": [
        ("Scoops", "Vanilla Scoop", 80, "https://images.unsplash.com/photo-1570197788417-0e82375c9371?w=600&q=80"),
        ("Scoops", "Chocolate Scoop", 90, "https://images.unsplash.com/photo-1563805042-7684c019e1cb?w=600&q=80"),
        ("Sundaes", "Hot Chocolate Fudge", 220, "https://images.unsplash.com/photo-1501443762994-82bd5dace89a?w=600&q=80"),
        ("Sundaes", "Brownie Sundae", 260, "https://images.unsplash.com/photo-1488900128323-21503983a07e?w=600&q=80"),
    ],
}


SEED_OFFERS = [
    {
        "title": "20% OFF on Dine-In",
        "description": "Flat 20% off on total bill at Gallops Restaurant. Valid all day.",
        "image": "https://static.prod-images.emergentagent.com/jobs/96e3f558-0c2d-422f-809b-d051bc4aeb09/images/f53cfa62cad475d86a5e9c34857b260f5b8d98f8f873bdaf33b9131108b67f48.png",
        "outlet_name": "Gallops Restaurant",
        "valid_until": "2026-12-31",
    },
    {
        "title": "Buy 1 Get 1 on Coffee",
        "description": "Buy any large coffee and get a small one free at Gallops Cafe.",
        "image": "https://images.unsplash.com/photo-1509042239860-f550ce710b93?w=1200&q=80",
        "outlet_name": "Gallops Cafe",
        "valid_until": "2026-12-31",
    },
    {
        "title": "Plaza-Wide Happy Hours",
        "description": "15% off across all outlets between 4-6 PM.",
        "image": "https://images.unsplash.com/photo-1555507036-ab1f4038808a?w=1200&q=80",
        "outlet_name": None,
        "valid_until": "2026-12-31",
    },
]


@app.on_event("startup")
async def startup():
    # Indexes
    await db.users.create_index("email", unique=True)
    await db.outlets.create_index("id", unique=True)
    await db.menu_items.create_index("id", unique=True)
    await db.menu_items.create_index("outlet_id")
    await db.offers.create_index("id", unique=True)
    await db.reservations.create_index("id", unique=True)
    await db.offer_claims.create_index("id", unique=True)
    await db.offer_claims.create_index("token", unique=True)
    await db.offer_claims.create_index("mobile")
    await db.offer_claims.create_index("created_at")

    # Seed admin
    existing_admin = await db.users.find_one({"email": ADMIN_EMAIL})
    if not existing_admin:
        await db.users.insert_one({
            "id": str(uuid.uuid4()),
            "email": ADMIN_EMAIL,
            "password_hash": hash_password(ADMIN_PASSWORD),
            "name": "Admin",
            "role": "admin",
            "created_at": datetime.now(timezone.utc),
        })
        logger.info("Seeded admin user: %s", ADMIN_EMAIL)
    # NOTE: we intentionally do NOT reset the admin password on every startup,
    # so password changes made via the /auth/change-password endpoint persist.

    # Seed outlets if empty
    outlet_count = await db.outlets.count_documents({})
    if outlet_count == 0:
        name_to_id: dict = {}
        for o in SEED_OUTLETS:
            outlet = Outlet(**o)
            await db.outlets.insert_one(outlet.dict())
            name_to_id[outlet.name] = outlet.id
        logger.info("Seeded %d outlets", len(SEED_OUTLETS))

        # Seed menu
        menu_docs = []
        for outlet_name, items in SEED_MENU_BY_OUTLET.items():
            outlet_id = name_to_id.get(outlet_name)
            if not outlet_id:
                continue
            for category, name, price, image in items:
                item = MenuItem(
                    outlet_id=outlet_id,
                    category=category,
                    name=name,
                    price=float(price),
                    image=image,
                )
                menu_docs.append(item.dict())
        if menu_docs:
            await db.menu_items.insert_many(menu_docs)
            logger.info("Seeded %d menu items", len(menu_docs))

        # Seed offers
        for o in SEED_OFFERS:
            outlet_id = name_to_id.get(o["outlet_name"]) if o.get("outlet_name") else None
            offer = Offer(
                title=o["title"],
                description=o["description"],
                image=o["image"],
                outlet_id=outlet_id,
                valid_until=o.get("valid_until"),
            )
            await db.offers.insert_one(offer.dict())
        logger.info("Seeded %d offers", len(SEED_OFFERS))


@app.on_event("shutdown")
async def shutdown():
    client.close()
